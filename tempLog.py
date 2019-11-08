#!/usr/bin/env python
import os, datetime
from openpyxl import Workbook, load_workbook

#store the current date and time from the system
time = datetime.datetime.now()

#use vcgenmod to get the current temp measurement, strip unnecessary formatting, convert temp from C to F
temp = float(os.popen('vcgencmd measure_temp').readline().strip().replace('temp=','').replace("'C",""))
temp = temp * 1.8 + 32

#open existing spreadsheet to log temp
sourceBook = load_workbook(filename = '/path/to/file')

sourceSheet = sourceBook.active

#record the time and temp reading
sourceSheet.cell(row = sourceSheet.max_row + 1, column = 1).value = time

sourceSheet.cell(row = sourceSheet.max_row, column = 2).value = temp

#save the spreadsheet
sourceBook.save('/path/to/file')
